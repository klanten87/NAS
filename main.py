from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook('AS1.xlsx')      #Laddar dokument
AS1 = wb['AS1']                     #Laddar flik AS1
rows = AS1.max_row                  #Kollar vilken sista raden ar
E = AS1.cell(row=2, column=5).value #Laddar vilka skript som ska köras.
Skyltlista = wb['Skyltlista']       # Laddar flik Skyltlista
Listor = wb['Listor']               # Laddar flik Listor
Avrop = wb['Avrop']                 # Laddar flik Avrop

def systemName10(rows):       #KLAR
    system = "Not defined"
    for i in range(7,rows):
        bold = AS1.cell(row=i, column=2).font
        beteckning = AS1.cell(row=i, column=2).value
        if bold.b is True:
            system =  AS1.cell(row=i, column=2).value
        elif beteckning is not None and system not in beteckning:
            AS1.cell(row=i, column=2).value = system + "-" + beteckning


def SkrivSkylt10(rows): #KLAR
    skylt = 27  # Startrad for skyltlista
    nosign =['Elmätare','Tidkanal','Larm']
    for i in range(7,rows): #Går egenom alla rader från rad 7 till sista raden.
        keyword = AS1.cell(row=i, column=3).value
        A = AS1.cell(row=i, column=1).value  #Plockar fram om någon skrivit något i Optionkolumn.
        C = AS1.cell(row=i, column=3).value
        D = AS1.cell(row=i, column=4).value
        for col in range(9,18): #Kollar igenom kolumn I-Q om det är någon I/O-kopplad till komponenten.
            if keyword is not None and any(x in keyword for x in nosign):
                break
            V = AS1.cell(row=i, column=col).value
            if V is not None:
                writeSign(i,skylt,"200")
                skylt += 1
                break
            if C is not None and "Rökdetektor" in C:
                writeSign(i, skylt, "200")
                skylt += 1
                break
        if D is not None and "Siox" in D:
            B = AS1.cell(row=i, column=2).value
            split = B.split("-")
            del split[-1]
            S = "-".join(split) + "-BSC" + str(AS1.cell(row=i, column=8).value)
            Skyltlista.cell(row=skylt, column=2).value = "TYP 200"
            Skyltlista.cell(row=skylt, column=8).value = "1"
            Skyltlista.cell(row=skylt, column=10).value = S
            Skyltlista.cell(row=skylt, column=11).value = "BRANDSPJÄLLSCENTRAL"
            Skyltlista.cell(row=skylt, column=12).value = AS1.cell(row=5, column=2).value.upper()
            skylt += 1
        if A is not None and "s" in A:
            writeSign(i, skylt, "100")
            skylt += 1


def writeSign(i, skylt, typ):   #KLAR tillhör SkrivSkylt10
    Skyltlista.cell(row=skylt, column=2).value = "TYP " + typ
    Skyltlista.cell(row=skylt, column=8).value = "1"
    Skyltlista.cell(row=skylt, column=10).value = AS1.cell(row=i, column=2).value.upper()
    Skyltlista.cell(row=skylt, column=11).value = AS1.cell(row=i, column=3).value.upper()
    if "200" in typ:
        Skyltlista.cell(row=skylt, column=12).value = AS1.cell(row=5, column=2).value.upper()


def SkrivMotor10(rows): #KLAR
    rad = 5  # Startrad for Motordata
    Motor = wb['Provning motorer']  # Laddar flik Provning motorer
    motor=['Frånluftsfläkt','Pump','Tilluftsfläkt']
    for i in range(7,rows):
        keyword = AS1.cell(row=i, column=3).value
        if keyword is not None and any(x in keyword for x in motor):
            F = AS1.cell(row=i, column=6).value
            Motor.cell(row=rad, column=1).value = AS1.cell(row=i, column=2).value
            Motor.cell(row=rad, column=2).value = AS1.cell(row=i, column=3).value
            F = F.split('@')
            Motor.cell(row=rad, column=5).value = F[-1]
            rad += 1


def SkrivEgenprovning10(rows):  #KLAR
    Egen = wb['Egenkontroll']  # Laddar flik Egenkontroll
    rad = 4 #Startrad
    switch =['Larm','Tidkanal','Elmätare']
    for i in range(7,rows):
        system = AS1.cell(row=i, column=2).font
        if system.b is True:
            rad += 1
            Egen.cell(row=rad, column=1).value = AS1.cell(row=i, column=2).value
            Egen.cell(row=rad, column=1).font = Font(bold=True)
            rad += 1
            continue
        B = AS1.cell(row=i, column=2).value
        C = AS1.cell(row=i, column=3).value
        if B is not None and C is not None and any(x in C for x in switch):
            E = AS1.cell(row=i, column=5).value
            if E is None:
                Egen.cell(row=rad, column=1).value = AS1.cell(row=i, column=2).value
                Egen.cell(row=rad, column=2).value = AS1.cell(row=i, column=3).value
                rad += 1
                continue
            Egen.cell(row=rad, column=1).value = AS1.cell(row=i, column=2).value
            Egen.cell(row=rad, column=2).value = AS1.cell(row=i, column=5).value
            rad += 1
        elif B is not None:
            Egen.cell(row=rad, column=1).value = AS1.cell(row=i, column=2).value
            Egen.cell(row=rad, column=2).value = AS1.cell(row=i, column=3).value
            rad += 1


def keywordSort10(rows):
    #Switch-funktion som kollar vad vi har och skickar sedan till rätt funktion.
    sort = {
        "Avluftsspjäll": Avluftsspjall,
        "Avluftstemperatur": Avluftstemperatur,
        "Brandspjäll": Brandspjall,
        "Differenstryckgivare": Differenstryckgivare,
        "Elmätare": Elmatare,
        "Expansionskärl": Expansionskarl,
        "Filtervakt": Filtervakt,
        "Flödesgivare": Flodesgivare,
        "Framledningstemperatur": Framledningstemperatur,
        "Frysskyddstemperatur": Frysskyddstemperatur,
        "Frånluftsfläkt": Franluftsflakt,
        "Frånluftsspjäll": Franluftsspjall,
        "Frånluftstemperatur": Franluftstemperatur,
        "Kallvattenmätare": Kallvattenmatare,
        "Kylmaskin": Kylmaskin,
        "Larm": Larm,
        "Ljusgivare": Ljusgivare,
        "Luftkvalitégivare": Luftkvalitegivare,
        "Pump": Pump,
        "Returledningstemperatur": Returledningstemperatur,
        "Rumstemperatur": Rumstemperatur,
        "Rökdetektor": Rokdetektor,
        "Spjällställdon": Spjallstalldon,
        "Tidkanal": Tidkanal,
        "Tilluftsfläkt": Tilluftsflakt,
        "Tilluftsspjäll": Tilluftsspjall,
        "Tilluftstemperatur": Tilluftstemperatur,
        "Tilluftstemperature.VVX": TilluftstemperaturVVX,
        "Tryckgivare": Tryckgivare,
        "Tryckvakt": Tryckvakt,
        "Uteluftsspjäll": Uteluftsspjall,
        "Uteluftskanalstemperatur": Uteluftskanalstemperatur,
        "Utomhustemperatur": Utomhustemperatur,
        "Varmvattenmätare": Varmvattenmatare,
        "VAV-spjäll": VAVspjall,
        "Ventilställdon": Ventilstalldon,
        "VVC-temperatur": VVCtemperatur,
        "Värmemängdsmätare": Varmemangdsmatare,
        "Värmeväxlare": Varmevaxlare,
        "Förlängdventilation": Forlangdventilation,
        "Serviceomkopplare": Serviceomkopplare,
        "Elbatteri": Elbatteri,
        "Fläktvakt": Flaktvakt,
    }
    material = 1
    best = 1
    larm = 1
    inst = 1
    system = "not Set"
    for i in range(7,rows):
        system = AS1.cell(row=i, column=2).font
        if system.b is True:
            material += 1
            larm += 1
            inst += 1
            system = AS1.cell(row=i, column=2).value
            Listor.cell(row=material, column=7).value = system
            Listor.cell(row=material, column=7).font = Font(bold=True)
            Listor.cell(row=larm, column=1).value = system
            Listor.cell(row=larm, column=1).font = Font(bold=True)
            Listor.cell(row=inst, column=12).value = system
            Listor.cell(row=inst, column=12).font = Font(bold=True)
            material += 1
            larm += 1
            inst += 1


        C = AS1.cell(row=i, column=3).value
        CBold = AS1.cell(row=i, column=3).font
        if C is not None and not "Reserv" in C and CBold.b is not True:
            if " " in C:
                C = C.split(" ")
                C = "".join(C)
            row_add = sort[C](i,material,larm,inst)
            row_add = list(row_add)
            larm += int(row_add[0])
            material += int(row_add[1])
            inst += int(row_add[2])


def Avluftsspjall(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Avluftstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    return("110")


def Brandspjall(i,material,larm,inst):
    #Material
    #Larm - Brandspjäll fel läge 2 m
    #Inställning - Motionering Söndag 23:55-23:59
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Brandspjäll i fel läge", "", "B", "5 m")
    skrivInst(i, inst, "Motionering av brandspjäll", "Söndag 23:55-23:59")
    return("111")


def Differenstryckgivare(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Regleravvikelse +/-3KPa 15 m
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Regleravvikelse", "+/-3 kPa", "C", "60 m")
    skrivInst(i, inst, "Börvärde differenstryck", "50 kPa")
    return("211")


def Elmatare(i,material,larm,inst):
    #Material
    skrivMaterial(i, material)
    return("010")


def Expansionskarl(i,material,larm,inst):
    #Larm - Låg tryck expansionkälr 2 m
    skrivLarm(i, larm, "Lågt tryck expansionskärl", "1 bar", "B", "2 m")
    return("100")


def Filtervakt(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Smutsiga filter 15 m
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Smutsiga filter", "70 Pa", "B", "2 m")
    return("110")


def Flodesgivare(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Regleravvikelse 15 l/s 15 m
    #Inställning - Börvärde flöde Min/Max	15/30 l/s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Regleravvikelse", "+/-15 l/s", "C", "60 m")
    skrivInst(i, inst, "Börvärde flöde Min/Max", "15/30 l/s")
    return("211")


def Framledningstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Regleravvikelse +/-3°C 15 m
    #Inställning - Börvärde Utetemp X -20°C, -10°C, 0°C, 10°C, 20°C
    #Inställning - Börvärde framledning 65°C, 55°C, 47°, 32°C, 20°C
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Regleravvikelse", "+/-3°C", "C", "60 m")
    skrivInst(i, inst, "Börvärde Utetemp X", "-20°C, -10°C, 0°C, 10°C, 20°C")
    inst += 1
    skrivInst(i, inst, "Börvärde framledning Y", "65°C, 55°C, 47°, 32°C, 20°C")
    return("212")


def Frysskyddstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Frysvaktslarm 7°C A 3 s
    #Inställning - Varmhållning vid stillastående 20°C
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Frysskyddslarm", "7°C", "A", "3 s")
    skrivInst(i, inst, "Varmhållning vid stillastående", "20°C")
    return("211")


def Franluftsflakt(i,material,larm,inst):
    #Material
    #Larm - Driftfel frånluftsfläkt 2 m
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Driftfel", "", "B", "2 m")
    return("110")


def Franluftsspjall(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Franluftstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    return("110")


def Kallvattenmatare(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Kylmaskin(i,material,larm,inst):
    #Material
    #Larm - Summalarm Kylmaskin 30 s
    skrivLarm(i, larm, "Summalarm kylmaskin", "", "B", "10 s")
    return("100")


def Larm(i,material,larm,inst):
    #Larm - %%Benämning%% 10 s
    text = AS1.cell(row=i, column=5).value
    skrivLarm(i,larm,text,"","B","10 s")
    return("100")


def Ljusgivare(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel
    #Inställning - Ljusnivå 80 Lux
    skrivMaterial(i,material)
    skrivLarm(i,larm,"Givarfel","","B","30 s")
    skrivInst(i,inst,"Ljusnivå dag","80 lux")
    return("111")


def Luftkvalitegivare(i,material,larm,inst):
    #Material
    #BEställning
    #Larm - Givarfel 30 s
    #Larm - Hög CO²-halt 900ppm 15 m
    #Inställning - Börvärde Luftkvalité 800 ppm
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Hög CO²-halt", "900ppm", "B", "5 m")
    skrivInst(i, inst, "Börvärde luftkvalité", "800 ppm")
    return("211")


def Pump(i,material,larm,inst):
    #Material
    #Larm - Driftfel pump 2 m
    #Inställning - Pumpstart 7°C
    #Inställning - Pumpstopp 17°C
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Driftfel", "", "B", "2 m")
    skrivInst(i, inst, "Pumpstart om utetemp <", "7°C")
    inst += 1
    skrivInst(i, inst, "Pumpstopp om utetemp >", "17°C")
    return("112")


def Returledningstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    return("110")


def Rumstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Hög rumstemperatur 23°C 15m
    #Larm - Låg rumstemperatur 17°C 15m
    #Inställning - Börvärde rumstemperatur 21°C
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Hög rumstemperatur", "23°C", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Låg rumstemperatur","17°C", "B", "5 m")
    skrivInst(i, inst, "Börvärde rumstemperatur", "21°C")
    return("311")


def Rokdetektor(i,material,larm,inst):
    ##Material
    #Beställning
    #Larm - Utlöst rökdetektor 10 s
    #Larm - Servicelarm rökdetektor 10 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Utlöst rökdetektor", "", "A", "10 s")
    larm += 1
    skrivLarm(i, larm, "Servicelarm rökdetektor", "", "B", "10s")
    return("210")


def Spjallstalldon(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Tidkanal(i,material,larm,inst):
    #Inställning - %%Benämning%% M-F 07:00-16:00
    text = AS1.cell(row=i, column=5).value
    skrivInst(i, inst, text, "M-F 07:00-16:00")
    return("001")


def Tilluftsflakt(i,material,larm,inst):
    #Material
    #Larm - Driftfel tilluftsfläkt 2 m
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Driftfel", "", "B", "2 m")
    return("110")


def Tilluftsspjall(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Tilluftstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Regleravvikelse +/-3°C 15m
    #Inställning - Börvärde tilluftstemp 19°C
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Regleravvikelse", "+/-3°C", "B", "60 m")
    skrivInst(i, inst, "Börvärde tilluftstemperatur", "19°C")
    return("211")


def TilluftstemperaturVVX(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    return("110")


def Tryckgivare(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Regleravvikelse +/-10Pa 15m
    #Larm - Fläktvakt 50Pa 5m
    #Inställning - Börvärde tryck 150 Pa
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Regleravvikelse", "+/-15 Pa", "B", "60 m")
    larm += 1
    skrivLarm(i, larm, "Fläktvakt", "<50 Pa", "B", "15 m")
    skrivInst(i, inst, "Tryckbörvärde", "150 Pa")
    return("311")


def Tryckvakt(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Driftfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Driftfel", "30 Pa", "B", "2 m")
    return("110")


def Flaktvakt(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Driftfel", "30 Pa", "B", "2 m")
    #Larm - Driftfel 30 s
    return("110")


def Uteluftsspjall(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Uteluftskanalstemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    return("110")


def Utomhustemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    return("110")


def Varmvattenmatare(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def VAVspjall(i,material,larm,inst):
    #Material
    skrivMaterial(i, material)
    return("010")


def Ventilstalldon(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def VVCtemperatur(i,material,larm,inst):
    #Material
    #Beställning
    #Larm - Givarfel 30 s
    #Larm - Låg returtempertur 40°C 15 m
    skrivMaterial(i, material)
    skrivLarm(i, larm, "Givarfel", "", "B", "30 s")
    larm += 1
    skrivLarm(i, larm, "Låg returtempertur", "<40°C", "B", "15 m")
    return("210")


def Varmemangdsmatare(i,material,larm,inst):
    #Material
    #Beställning
    skrivMaterial(i, material)
    return("010")


def Varmevaxlare(i,material,larm,inst):
    #Larm - Summalarm 10s
    #Larm - Låg verkningsgrad 50% 30 m
    skrivLarm(i, larm, "Summalarm", "", "B", "10 s")
    larm += 1
    skrivLarm(i, larm, "Låg verkningsgrad", "<50%", "B", "30 m")
    return("200")


def Forlangdventilation(i,material,larm,inst):
    #Material
    #Beställning
    #Instälning - Eftergångstid 1 timme
    skrivMaterial(i, material)
    skrivInst(i, inst, "Eftergångstid", "1 timme")
    return("011")


def Serviceomkopplare(i,material,larm,inst):
    #Larm - Serviceomkopplare i fel läge 10s
    skrivLarm(i, larm, "Serviceomkopplare i fel läge", "", "B", "10 s")
    return("100")


def Elbatteri(i,material,larm,inst):
    #Material
    #Larm - Överhettningslarm
    #Inställningsvärde - Eftergångstid 5m
    skrivMaterial(i, material)
    skrivInst(i, inst, "Efterblåsning tilluft", "5 m")
    return("011")


def skrivInst(i,inst,obj,levinst):
    Listor.cell(row=inst, column=12).value = AS1.cell(row=i, column=2).value
    Listor.cell(row=inst, column=13).value = obj
    Listor.cell(row=inst, column=14).value = levinst


def skrivLarm(i,larm,text,level,klass,delay):
    Listor.cell(row=larm, column=1).value = AS1.cell(row=i, column=2).value
    Listor.cell(row=larm, column=2).value = text
    Listor.cell(row=larm, column=3).value = klass
    Listor.cell(row=larm, column=4).value = level
    Listor.cell(row=larm, column=5).value = delay


def skrivMaterial(i,material):
    fabrikat = AS1.cell(row=i, column=4).value
    typ = AS1.cell(row=i, column=5).value
    Listor.cell(row=material, column=7).value = AS1.cell(row=i, column=2).value
    Listor.cell(row=material, column=8).value = AS1.cell(row=i, column=3).value
    if fabrikat is not None and typ is not None and fabrikat in typ:
        Listor.cell(row=material, column=9).value = typ
    elif fabrikat is not None and typ is not None:
        Listor.cell(row=material, column=9).value = fabrikat + " " + typ
    else:
        Listor.cell(row=material, column=9).value = typ


def skrivAvrop(rows):
    fabrikat = [
         "Abelko",
         "Belimo",
         "Calectro",
         "Danfoss",
         "EkoVent",
         "ESBE",
         "Fidelix",
         "Fläktwoods",
         "Grundfos",
         "HAGAB",
         "IV",
         "Kamstrup",
         "Produal",
         "Regin",
         "Schnieder",
         "Siemens",
         "Siox",
         "Swegon"
    ]
    avrop = 1
    for f in fabrikat:
        rubrik = 0
        for i in range(7, rows):
            D = AS1.cell(row=i, column=4).value
            if D is not None and f in D:
                if rubrik is 0:
                    avrop += 1
                    Avrop.cell(row=avrop, column=1).value = f
                    Avrop.cell(row=avrop, column=1).font = Font(bold=True)
                    rubrik = 1
                    avrop += 1
                E = AS1.cell(row=i, column=5).value
                if E is not None and "," in E:
                    E = E.split(",")
                    for e in E:
                        Avrop.cell(row=avrop, column=1).value = e
                        Avrop.cell(row=avrop, column=2).value = AS1.cell(row=i, column=2).value
                        Avrop.cell(row=avrop, column=3).value = "1"
                        avrop += 1
                else:
                    Avrop.cell(row=avrop, column=1).value = E
                    Avrop.cell(row=avrop, column=2).value = AS1.cell(row=i, column=2).value
                    Avrop.cell(row=avrop, column=3).value = "1"
                    avrop +=1




E = E.upper()
if "L" in E:
    keywordSort10(rows)
if "B" in E:
    systemName10(rows)
if "S" in E:
    SkrivSkylt10(rows)
if "M" in E:
    SkrivMotor10(rows)
if "E" in E:
    SkrivEgenprovning10(rows)
if "A" in E:
    skrivAvrop(rows)

wb.save('AS1_genererad.xlsx')
print("Done")